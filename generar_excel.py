#!/usr/bin/env python3
"""
Genera un Excel visual desde el CSV enriquecido.
"""

import csv
import sys
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.worksheet.filters import AutoFilter

INPUT_FILE = "base_holistica_20260529_0832_enriquecido.csv"
OUTPUT_FILE = "base_holistica_visual.xlsx"

# ── Paleta ──────────────────────────────────────────────────────────────
COL_HEADER_BG   = "1F4E79"   # azul oscuro
COL_HEADER_FG   = "FFFFFF"
COL_SCORE_HIGH  = "C6EFCE"   # verde claro (score >= 7)
COL_SCORE_MID   = "FFEB9C"   # amarillo (score 4-6)
COL_SCORE_LOW   = "FFC7CE"   # rojo claro (score <= 3)
COL_ROW_ALT     = "EBF3FB"   # azul muy claro para filas alternas
COL_HAS_EMAIL   = "E2EFDA"   # verde suave si tiene email
COL_NO_EMAIL    = "FCE4D6"   # naranja suave si NO tiene email
COL_BORDER      = "BDD7EE"

COLUMN_WIDTHS = {
    "empresa": 32,
    "rubro": 24,
    "direccion": 36,
    "ciudad": 18,
    "provincia": 18,
    "telefono": 16,
    "website": 30,
    "google_maps_url": 12,
    "rating": 8,
    "reviews_count": 10,
    "price_level": 10,
    "horarios": 30,
    "instagram": 20,
    "email": 28,
    "whatsapp": 16,
    "score_ia": 10,
    "observaciones": 24,
    "fuente": 18,
    "fecha_extraccion": 16,
}

FRIENDLY_HEADERS = {
    "empresa": "Empresa",
    "rubro": "Rubro",
    "direccion": "Dirección",
    "ciudad": "Ciudad",
    "provincia": "Provincia",
    "telefono": "Teléfono",
    "website": "Website",
    "google_maps_url": "Maps",
    "rating": "Rating ⭐",
    "reviews_count": "Reviews",
    "price_level": "Precio $",
    "horarios": "Horarios",
    "instagram": "Instagram",
    "email": "Email",
    "whatsapp": "WhatsApp",
    "score_ia": "Score IA",
    "observaciones": "Observaciones",
    "fuente": "Fuente",
    "fecha_extraccion": "Fecha",
}

def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    side = Side(style="thin", color=COL_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)

def score_fill(score_str):
    try:
        s = float(score_str)
    except (ValueError, TypeError):
        return None
    if s >= 7:
        return hex_fill(COL_SCORE_HIGH)
    elif s >= 4:
        return hex_fill(COL_SCORE_MID)
    else:
        return hex_fill(COL_SCORE_LOW)

def build_excel(input_file, output_file):
    with open(input_file, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames)
        rows = list(reader)

    wb = Workbook()

    # ── Hoja 1: Base Completa ───────────────────────────────────────────
    ws = wb.active
    ws.title = "Base Completa"
    ws.freeze_panes = "A2"

    # Encabezados
    for col_idx, field in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=FRIENDLY_HEADERS.get(field, field))
        cell.fill = hex_fill(COL_HEADER_BG)
        cell.font = Font(bold=True, color=COL_HEADER_FG, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = thin_border()

    ws.row_dimensions[1].height = 22

    score_col = fieldnames.index("score_ia") + 1 if "score_ia" in fieldnames else None
    email_col = fieldnames.index("email") + 1 if "email" in fieldnames else None

    for row_idx, row in enumerate(rows, 2):
        alt = (row_idx % 2 == 0)
        for col_idx, field in enumerate(fieldnames, 1):
            val = row.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = thin_border()

            # Color base (filas alternas)
            if alt:
                cell.fill = hex_fill(COL_ROW_ALT)

            # Score IA → color condicional en esa columna
            if col_idx == score_col:
                sf = score_fill(val)
                if sf:
                    cell.fill = sf
                    cell.font = Font(bold=True, size=9)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Email → verde si tiene, naranja si no
            if col_idx == email_col:
                cell.fill = hex_fill(COL_HAS_EMAIL) if val.strip() else hex_fill(COL_NO_EMAIL)

        ws.row_dimensions[row_idx].height = 16

    # Anchos de columna
    for col_idx, field in enumerate(fieldnames, 1):
        width = COLUMN_WIDTHS.get(field, 16)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Autofiltro
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}1"

    # ── Hoja 2: Solo leads con email ────────────────────────────────────
    ws2 = wb.create_sheet("Con Email")
    ws2.freeze_panes = "A2"
    leads_email = [r for r in rows if r.get("email", "").strip()]

    for col_idx, field in enumerate(fieldnames, 1):
        cell = ws2.cell(row=1, column=col_idx, value=FRIENDLY_HEADERS.get(field, field))
        cell.fill = hex_fill("1B6B3A")
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
    ws2.row_dimensions[1].height = 22

    for row_idx, row in enumerate(leads_email, 2):
        for col_idx, field in enumerate(fieldnames, 1):
            val = row.get(field, "")
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border()
            if row_idx % 2 == 0:
                cell.fill = hex_fill("EAF5EA")
            if col_idx == score_col:
                sf = score_fill(val)
                if sf:
                    cell.fill = sf
                    cell.font = Font(bold=True, size=9)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[row_idx].height = 16

    for col_idx, field in enumerate(fieldnames, 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(field, 16)
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}1"

    # ── Hoja 3: Score IA >= 7 ───────────────────────────────────────────
    ws3 = wb.create_sheet("Score Alto (≥7)")
    ws3.freeze_panes = "A2"
    leads_high = [r for r in rows if r.get("score_ia", "").strip() and float(r["score_ia"]) >= 7]

    for col_idx, field in enumerate(fieldnames, 1):
        cell = ws3.cell(row=1, column=col_idx, value=FRIENDLY_HEADERS.get(field, field))
        cell.fill = hex_fill("7B3F00")
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
    ws3.row_dimensions[1].height = 22

    for row_idx, row in enumerate(leads_high, 2):
        for col_idx, field in enumerate(fieldnames, 1):
            val = row.get(field, "")
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border()
            if row_idx % 2 == 0:
                cell.fill = hex_fill("FFF8EE")
            if col_idx == score_col:
                sf = score_fill(val)
                if sf:
                    cell.fill = sf
                    cell.font = Font(bold=True, size=9)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[row_idx].height = 16

    for col_idx, field in enumerate(fieldnames, 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(field, 16)
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}1"

    # ── Hoja 4: Resumen / Dashboard ─────────────────────────────────────
    ws4 = wb.create_sheet("Resumen")
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 14
    ws4.column_dimensions["C"].width = 10

    title = ws4.cell(row=1, column=1, value="BASE HOLÍSTICA ARGENTINA — RESUMEN")
    title.font = Font(bold=True, size=14, color="1F4E79")
    title.alignment = Alignment(horizontal="left")
    ws4.merge_cells("A1:C1")
    ws4.row_dimensions[1].height = 28

    stats = [
        ("Total registros", len(rows), ""),
        ("Con teléfono", sum(1 for r in rows if r.get("telefono","")), f"{100*sum(1 for r in rows if r.get('telefono',''))//max(len(rows),1)}%"),
        ("Con website", sum(1 for r in rows if r.get("website","")), f"{100*sum(1 for r in rows if r.get('website',''))//max(len(rows),1)}%"),
        ("Con email", sum(1 for r in rows if r.get("email","")), f"{100*sum(1 for r in rows if r.get('email',''))//max(len(rows),1)}%"),
        ("Con Instagram", sum(1 for r in rows if r.get("instagram","")), f"{100*sum(1 for r in rows if r.get('instagram',''))//max(len(rows),1)}%"),
        ("Score IA ≥ 7", sum(1 for r in rows if r.get("score_ia","").strip() and float(r["score_ia"]) >= 7), f"{100*sum(1 for r in rows if r.get('score_ia','').strip() and float(r['score_ia']) >= 7)//max(len(rows),1)}%"),
        ("Score IA ≥ 5", sum(1 for r in rows if r.get("score_ia","").strip() and float(r["score_ia"]) >= 5), f"{100*sum(1 for r in rows if r.get('score_ia','').strip() and float(r['score_ia']) >= 5)//max(len(rows),1)}%"),
    ]

    label_colors = ["1F4E79","2E75B6","2E75B6","1B6B3A","7030A0","7B3F00","C65911"]

    for i, (label, value, pct) in enumerate(stats, 3):
        lc = ws4.cell(row=i, column=1, value=label)
        vc = ws4.cell(row=i, column=2, value=value)
        pc = ws4.cell(row=i, column=3, value=pct)

        color = label_colors[i - 3] if (i - 3) < len(label_colors) else "333333"
        lc.font = Font(bold=(i == 3), size=11, color=color)
        vc.font = Font(bold=True, size=13, color=color)
        pc.font = Font(size=10, color="666666")
        vc.alignment = Alignment(horizontal="center")
        pc.alignment = Alignment(horizontal="center")

        if i == 3:
            lc.fill = hex_fill("DEEAF1")
            vc.fill = hex_fill("DEEAF1")
            pc.fill = hex_fill("DEEAF1")

        ws4.row_dimensions[i].height = 22

    wb.save(output_file)
    print(f"✅ Excel generado: {output_file}")
    print(f"   Hojas: Base Completa ({len(rows)}), Con Email ({len(leads_email)}), Score Alto ({len(leads_high)}), Resumen")

if __name__ == "__main__":
    build_excel(INPUT_FILE, OUTPUT_FILE)
